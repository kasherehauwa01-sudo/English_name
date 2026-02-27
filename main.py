#!/usr/bin/env python3
"""Поиск латиницы в Excel и формирование итогового XLSX-отчёта."""

from __future__ import annotations

import argparse
import logging
import re
import sys
from dataclasses import dataclass
from io import BytesIO
from pathlib import Path
from typing import Optional

import pandas as pd

LATIN_RE = re.compile(r"[A-Za-z]")
LATIN_TOKEN_RE = re.compile(r"[A-Za-z0-9]+(?:-[A-Za-z0-9]+)*")
VOWELS = set("aeiouy")

EXCEPTION_MAP = {
    "usb": "усб",
    "led": "лед",
    "wifi": "вайфай",
    "wi-fi": "вай-фай",
    "bluetooth": "блютус",
    "smart": "смарт",
    "pro": "про",
    "max": "макс",
    "mini": "мини",
    "ultra": "ультра",
    "eco": "эко",
    "iphone": "айфон",
    "samsung": "самсунг",
    "tefal": "тефаль",
    "type-c": "тайп-си",
    "usb-c": "усб-си",
}

TRANSLATE_MAP = {
    "smart": "умный",
    "wireless": "беспроводной",
    "portable": "портативный",
    "mini": "мини",
    "max": "максимум",
    "ultra": "ультра",
    "pro": "профессиональный",
    "light": "свет",
    "lamp": "лампа",
    "router": "роутер",
    "speaker": "колонка",
    "watch": "часы",
    "phone": "телефон",
    "case": "чехол",
    "cable": "кабель",
    "charger": "зарядное устройство",
}


@dataclass
class MatchRow:
    source_file: str
    code: str
    name: str
    transcription: str


def setup_logging(log_file: str = "parse.log") -> None:
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s | %(levelname)s | %(name)s | %(message)s",
        handlers=[logging.StreamHandler(sys.stdout), logging.FileHandler(log_file, mode="w", encoding="utf-8")],
    )


def normalize_header(value: object) -> str:
    text = str(value or "").replace("\r", " ").replace("\n", " ")
    text = text.strip().lower().replace("ё", "е")
    return re.sub(r"\s+", " ", text)


def find_required_columns(columns: list[object]) -> tuple[Optional[object], Optional[object]]:
    code_col = None
    name_col = None

    for col in columns:
        if normalize_header(col) == "код":
            code_col = col
            break

    for col in columns:
        n = normalize_header(col)
        if name_col is None and n in {"наименование товаров", "наименование товара", "наименование"}:
            name_col = col

    if code_col is None:
        for col in columns:
            n = normalize_header(col)
            if n in {"код товара", "артикул"}:
                code_col = col
                break

    if name_col is None:
        for col in columns:
            n = normalize_header(col)
            if "наимен" in n and "товар" in n:
                name_col = col
                break

    return code_col, name_col


def letter_translit(chunk: str) -> str:
    s = chunk.lower()
    for old, new in [("sch", "щ"), ("sh", "ш"), ("ch", "ч"), ("zh", "ж"), ("kh", "х"), ("ph", "ф"), ("th", "т"), ("ck", "к"), ("qu", "кв")]:
        s = s.replace(old, new)

    out: list[str] = []
    for i, c in enumerate(s):
        nxt = s[i + 1] if i + 1 < len(s) else ""
        prv = s[i - 1] if i > 0 else ""
        mapping = {
            "a": "а", "b": "б", "d": "д", "e": "е", "f": "ф", "h": "х", "i": "и",
            "j": "дж", "k": "к", "l": "л", "m": "м", "n": "н", "o": "о", "p": "п",
            "q": "кв", "r": "р", "t": "т", "u": "у", "v": "в", "w": "в", "x": "кс", "z": "з",
        }
        if c.isdigit():
            out.append(c)
        elif c == "c":
            out.append("с" if nxt in "eiy" else "к")
        elif c == "g":
            out.append("дж" if nxt in "eiy" else "г")
        elif c == "s":
            out.append("з" if prv in VOWELS and nxt in VOWELS else "с")
        elif c == "y":
            out.append("й" if i == 0 else "и")
        elif c == "-":
            out.append("-")
        else:
            out.append(mapping.get(c, c))
    return "".join(out)


def apply_case_style(source_token: str, transliterated: str) -> str:
    if source_token.isupper():
        return transliterated.upper()
    if source_token.islower():
        return transliterated.lower()
    if source_token[:1].isupper() and source_token[1:].islower():
        return transliterated[:1].upper() + transliterated[1:].lower() if transliterated else transliterated
    return transliterated


def translit_token(token: str) -> str:
    if re.fullmatch(r"[A-Za-z]{1,2}", token):
        return token

    low = token.lower()
    if low in EXCEPTION_MAP:
        return apply_case_style(token, EXCEPTION_MAP[low])

    parts = token.split("-")
    if len(parts) > 1:
        return "-".join(translit_token(p) for p in parts)

    mixed = re.findall(r"[A-Za-z]+|\d+", token)
    if len(mixed) > 1:
        return "".join(p if p.isdigit() else translit_token(p) for p in mixed)

    return apply_case_style(token, letter_translit(token))


def translit_to_ru(text: str) -> str:
    return LATIN_TOKEN_RE.sub(lambda m: translit_token(m.group(0)), text)


def deduplicate_rows_by_code(rows: list[MatchRow]) -> list[MatchRow]:
    seen: set[str] = set()
    out: list[MatchRow] = []
    for row in rows:
        key = row.code.strip()
        if key in seen:
            continue
        seen.add(key)
        out.append(row)
    return out




def translate_name_to_ru(name: str) -> tuple[str, bool]:
    """Переводит переводимые английские слова в названии на русский аналог."""
    replaced = False

    def repl(match: re.Match[str]) -> str:
        nonlocal replaced
        token = match.group(0)
        low = token.lower()
        if low in TRANSLATE_MAP:
            replaced = True
            return TRANSLATE_MAP[low]
        return token

    translated = LATIN_TOKEN_RE.sub(repl, name)
    return translated, replaced


def build_translatable_rows(rows: list[MatchRow]) -> list[dict[str, str]]:
    """Формирует строки для вкладки 'Переводимые'."""
    out: list[dict[str, str]] = []
    for r in rows:
        translated, has_translation = translate_name_to_ru(r.name)
        if not has_translation:
            continue
        out.append(
            {
                "Код": r.code,
                "Наименование товара": r.name,
                "Перевод": translated,
                "Файл источник": r.source_file,
            }
        )
    return out


def write_xlsx(rows: list[MatchRow], out_file: str) -> str:
    final = out_file if out_file.lower().endswith(".xlsx") else f"{out_file}.xlsx"

    latin_data = [
        {
            "Код": r.code,
            "Наименование товара": r.name,
            "Транскрипция": r.transcription,
            "Файл источник": r.source_file,
        }
        for r in rows
    ]
    latin_df = pd.DataFrame(latin_data)

    translated_data = build_translatable_rows(rows)
    translated_df = pd.DataFrame(translated_data, columns=["Код", "Наименование товара", "Перевод", "Файл источник"])

    with pd.ExcelWriter(final, engine="openpyxl") as writer:
        latin_df.to_excel(writer, index=False, sheet_name="latin")
        translated_df.to_excel(writer, index=False, sheet_name="Переводимые")
    return final


def _rows_with_merged_cells_xlsx(ws) -> set[int]:
    rows: set[int] = set()
    for mr in ws.merged_cells.ranges:
        rows.update(range(mr.min_row, mr.max_row + 1))
    return rows


def _rows_with_merged_cells_xls(sheet) -> set[int]:
    rows: set[int] = set()
    for rlow, rhigh, _clow, _chigh in getattr(sheet, "merged_cells", []):
        for r in range(rlow, rhigh):
            rows.add(r + 1)
    return rows


def _dataframe_from_xlsx_sheet(ws) -> pd.DataFrame:
    merged_rows = _rows_with_merged_cells_xlsx(ws)
    rows: list[list[object]] = []
    for r in range(1, ws.max_row + 1):
        if r <= 9 or r in merged_rows:
            continue
        values = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
        if any(v not in (None, "") for v in values):
            rows.append(values)
    if not rows:
        return pd.DataFrame()
    header = [str(h).strip() if h is not None else "" for h in rows[0]]
    return pd.DataFrame(rows[1:], columns=header)


def _dataframe_from_xls_sheet(sheet) -> pd.DataFrame:
    merged_rows = _rows_with_merged_cells_xls(sheet)
    rows: list[list[object]] = []
    for r in range(sheet.nrows):
        row_num = r + 1
        if row_num <= 9 or row_num in merged_rows:
            continue
        vals = sheet.row_values(r)
        if any(v not in (None, "") for v in vals):
            rows.append(vals)
    if not rows:
        return pd.DataFrame()
    header = [str(h).strip() if h is not None else "" for h in rows[0]]
    return pd.DataFrame(rows[1:], columns=header)


def _collect_rows_from_xlsx(source_name: str, binary: bytes | None = None, path: str | None = None) -> tuple[list[MatchRow], bool]:
    from openpyxl import load_workbook

    wb = load_workbook(filename=BytesIO(binary) if binary is not None else path, data_only=True)
    out: list[MatchRow] = []
    has_required = False
    for sheet_name in wb.sheetnames:
        df = _dataframe_from_xlsx_sheet(wb[sheet_name])
        if df.empty:
            continue
        code_col, name_col = find_required_columns(list(df.columns))
        if not code_col or not name_col:
            continue
        has_required = True
        names = df[name_col].fillna("").astype(str)
        codes = df[code_col].fillna("").astype(str)
        for code_value, name_value in zip(codes, names):
            name_text = name_value.strip()
            if name_text and LATIN_RE.search(name_text):
                out.append(MatchRow(source_name, str(code_value).strip(), name_text, translit_to_ru(name_text)))
    logging.getLogger("main").info("%s: найдено строк с латиницей: %s", source_name, len(out))
    return out, has_required


def _collect_rows_from_xls(source_name: str, binary: bytes | None = None, path: str | None = None) -> tuple[list[MatchRow], bool]:
    import xlrd

    book = xlrd.open_workbook(file_contents=binary, formatting_info=True) if binary is not None else xlrd.open_workbook(path, formatting_info=True)
    out: list[MatchRow] = []
    has_required = False
    for sheet in book.sheets():
        df = _dataframe_from_xls_sheet(sheet)
        if df.empty:
            continue
        code_col, name_col = find_required_columns(list(df.columns))
        if not code_col or not name_col:
            continue
        has_required = True
        names = df[name_col].fillna("").astype(str)
        codes = df[code_col].fillna("").astype(str)
        for code_value, name_value in zip(codes, names):
            name_text = name_value.strip()
            if name_text and LATIN_RE.search(name_text):
                out.append(MatchRow(source_name, str(code_value).strip(), name_text, translit_to_ru(name_text)))
    logging.getLogger("main").info("%s: найдено строк с латиницей: %s", source_name, len(out))
    return out, has_required


def _collect_rows_from_file(source_name: str, suffix: str, binary: bytes | None = None, path: str | None = None) -> tuple[list[MatchRow], bool]:
    if suffix.lower() == ".xlsx":
        return _collect_rows_from_xlsx(source_name, binary=binary, path=path)
    if suffix.lower() == ".xls":
        return _collect_rows_from_xls(source_name, binary=binary, path=path)
    raise ValueError(f"Неподдерживаемый формат файла: {suffix}")


def scan_uploaded_files(uploaded_files: list, out_file: str = "Готовый с транскрипцией.xlsx", status_callback=None, progress_callback=None) -> dict:
    logger = logging.getLogger("main")

    def set_status(msg: str) -> None:
        logger.info(msg)
        if status_callback:
            status_callback(msg)

    def set_progress(current: int, total: int) -> None:
        if progress_callback:
            progress_callback(current, total)

    total = len(uploaded_files)
    set_status(f"Загружено файлов: {total}")
    set_progress(0, total if total > 0 else 1)

    all_rows: list[MatchRow] = []
    errors: list[tuple[str, str]] = []
    files_with_columns = 0

    for idx, uf in enumerate(uploaded_files, start=1):
        set_progress(idx - 1, total)
        set_status(f"Обрабатываю файл {idx}/{total}: {uf.name}")
        try:
            rows, has_required = _collect_rows_from_file(uf.name, Path(uf.name).suffix, binary=uf.getvalue())
            all_rows.extend(rows)
            if has_required:
                files_with_columns += 1
        except Exception as exc:
            errors.append((uf.name, str(exc)))
            logger.error("Ошибка чтения %s: %s", uf.name, exc)

    set_progress(total, total if total > 0 else 1)
    set_status("Удаляю дубликаты по колонке 'Код'...")
    unique_rows = deduplicate_rows_by_code(all_rows)

    set_status("Формирую итоговый XLSX...")
    final_output = write_xlsx(unique_rows, out_file)
    set_status("Готово")

    return {
        "files": total,
        "files_with_columns": files_with_columns,
        "written": len(unique_rows),
        "errors": errors,
        "output": final_output,
    }


def scan_folder(folder_path: str, out_file: str = "Готовый с транскрипцией.xlsx", status_callback=None, progress_callback=None) -> dict:
    folder = Path(folder_path)
    uploaded_like = []
    for p in sorted([x for x in folder.iterdir() if x.is_file() and x.suffix.lower() in {'.xls','.xlsx'}]):
        class F: pass
        f=F(); f.name=p.name; f.getvalue=lambda p=p: p.read_bytes()
        uploaded_like.append(f)
    return scan_uploaded_files(uploaded_like, out_file, status_callback, progress_callback)


def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(description="Поиск латиницы в 'Наименование товаров' и экспорт в .xlsx")
    p.add_argument("--folder", required=True, help="Путь к папке с Excel-файлами")
    p.add_argument("--out", default="Готовый с транскрипцией.xlsx", help="Путь к итоговому .xlsx")
    return p.parse_args()


def main() -> None:
    args = parse_args()
    setup_logging()
    result = scan_folder(args.folder, args.out)
    logging.getLogger("main").info("Готово. Строк: %s | Файл: %s", result["written"], result["output"])


if __name__ == "__main__":
    main()
